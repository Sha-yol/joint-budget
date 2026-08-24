"""Auto-discover and classify input files in a directory."""

import csv
from pathlib import Path

import openpyxl


def _classify_csv(file_path: str) -> str | None:
    """
    Classify a CSV as 'splitwise_regular', 'splitwise_group', or None.

    Splitwise CSVs have a header starting with "Date,Description,Category,Cost,Currency".
    Regular (1:1) exports have exactly 2 person columns after Currency.
    Group exports have 3+ person columns.
    """
    with open(file_path, "r", encoding="utf-8") as f:
        lines = f.readlines()

    header_line = None
    for line in lines:
        stripped = line.strip()
        if stripped.startswith("Date,") or stripped.startswith("Date\t"):
            header_line = line
            break

    if header_line is None:
        return None

    delimiter = "\t" if "\t" in header_line else ","
    reader = csv.reader([header_line], delimiter=delimiter)
    fields = [f.strip() for f in next(reader)]

    expected_start = ["Date", "Description", "Category", "Cost", "Currency"]
    if fields[:5] != expected_start:
        return None

    person_cols = [f.strip() for f in fields[5:] if f.strip()]

    if len(person_cols) <= 2:
        return "splitwise_regular"
    else:
        return "splitwise_group"


def _xlsx_has_marker(file_path: str, marker: str, max_row: int, sheet: str | None = None) -> bool:
    """Return True if `marker` appears in column A of the first `max_row` rows.

    If `sheet` is given, that named sheet is used (returns False if missing);
    otherwise the active sheet is used.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if sheet is not None:
            if sheet not in wb.sheetnames:
                return False
            ws = wb[sheet]
        else:
            ws = wb.active
        for row in ws.iter_rows(max_row=max_row, values_only=True):
            cell = str(row[0] or "")
            if marker in cell:
                return True
        return False
    except Exception:
        return False


def discover_files(input_dir: str) -> dict:
    """
    Auto-discover and classify input files in a directory.

    Returns:
        {
            "mizrahi_ccs": [path, ...],
            "splitwise_regular": [path, ...],
            "splitwise_group": [path, ...],
        }
    """
    input_path = Path(input_dir)
    result = {"mizrahi_ccs": [], "splitwise_regular": [], "splitwise_group": []}

    for f in sorted(input_path.iterdir()):
        path = str(f)
        if f.suffix.lower() in (".xlsx", ".xls"):
            if _xlsx_has_marker(path, "חשבון כרטיס", max_row=3, sheet="חיובים בשקלים"):
                result["mizrahi_ccs"].append(path)
        elif f.suffix.lower() == ".csv":
            csv_type = _classify_csv(str(f))
            if csv_type:
                result[csv_type].append(str(f))

    return result
