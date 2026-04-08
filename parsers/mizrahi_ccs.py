"""Parser for Mizrahi bank's 'all credit cards' xlsx export."""

import re
from datetime import datetime

import openpyxl

from .common import make_expense


_NIS_SHEET = "חיובים בשקלים"
_CARD_HEADER_MARKER = "חשבון כרטיס"
_LAST4_RE = re.compile(r"ארבע ספרות אחרונות\s*(\d{4})")
_TABLE_HEADER = "תאריך חיוב"
_DATE_RE = re.compile(r"^\d{2}/\d{2}/\d{4}$")


def parse_mizrahi_ccs_excel(file_path: str) -> list[dict]:
    """
    Parse a Mizrahi bank 'all cards' xlsx export.

    The NIS sheet contains one or more stacked card blocks. Each block starts
    with a header cell `חשבון כרטיס: ... ארבע ספרות אחרונות NNNN`, followed by
    a blank row, the table header `תאריך חיוב | תאריך העסקה | בית העסק | סכום
    העסקה | סכום החיוב`, then transaction rows. Foreign-currency sheets exist
    but are currently empty and ignored.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if _NIS_SHEET not in wb.sheetnames:
        return []
    ws = wb[_NIS_SHEET]

    expenses: list[dict] = []
    current_last4: str | None = None
    in_table = False

    for row in ws.iter_rows(values_only=True):
        first = str(row[0] or "")

        if _CARD_HEADER_MARKER in first:
            m = _LAST4_RE.search(first)
            current_last4 = m.group(1) if m else None
            in_table = False
            continue

        if first.strip() == _TABLE_HEADER:
            in_table = True
            continue

        if not (in_table and current_last4):
            continue

        date_str = row[1]
        if not isinstance(date_str, str) or not _DATE_RE.match(date_str):
            continue

        try:
            formatted_date = datetime.strptime(date_str, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            continue

        raw_amount = row[4]
        try:
            amount = round(float(raw_amount), 2)
        except (TypeError, ValueError):
            continue
        if amount == 0:
            continue

        expenses.append(make_expense(
            date=formatted_date,
            source=current_last4,
            description=str(row[2] or "").strip(),
            amount=amount,
        ))

    return expenses
